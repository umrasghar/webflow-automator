# data/excel_integrator.py
# WebFlow Automator - Excel Integrator
# This module handles Excel file reading and writing

import os
import logging
import datetime
from typing import Dict, Any, List, Optional, Union, Tuple

# Try to import pandas and openpyxl
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from core.message_bus import MessageBus, MessageTypes

logger = logging.getLogger("WebFlowAutomator.Data.ExcelIntegrator")

class ExcelIntegrator:
    """
    Handles Excel file reading and writing for data import/export
    """
    
    def __init__(self, message_bus: MessageBus):
        """
        Initialize the Excel integrator
        
        Args:
            message_bus: Message bus for communication
        """
        self.message_bus = message_bus
        
        # Check if required libraries are available
        if not PANDAS_AVAILABLE:
            logger.warning("pandas library not available. Excel functionality will be limited.")
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl library not available. Excel functionality will be limited.")
    
    def can_read_excel(self) -> bool:
        """
        Check if Excel reading is supported
        
        Returns:
            bool: True if supported, False otherwise
        """
        return PANDAS_AVAILABLE or OPENPYXL_AVAILABLE
    
    def can_write_excel(self) -> bool:
        """
        Check if Excel writing is supported
        
        Returns:
            bool: True if supported, False otherwise
        """
        return PANDAS_AVAILABLE or OPENPYXL_AVAILABLE
    
    def read_excel(self, file_path: str, sheet_name: Optional[str] = None, 
                  start_row: int = 0, header: bool = True) -> List[Dict[str, Any]]:
        """
        Read data from an Excel file
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to read (None for first sheet)
            start_row: Row to start reading from (0-based)
            header: Whether the first row contains headers
        
        Returns:
            list: List of row data dictionaries
        """
        if not self.can_read_excel():
            raise RuntimeError("Excel reading functionality is not available. Please install pandas and/or openpyxl.")
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        try:
            if PANDAS_AVAILABLE:
                # Use pandas to read Excel
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=0 if header else None)
                else:
                    df = pd.read_excel(file_path, header=0 if header else None)
                
                # Skip rows if needed
                if start_row > 0:
                    df = df.iloc[start_row-1:]
                    # Reset index
                    df = df.reset_index(drop=True)
                
                # Convert to list of dictionaries
                result = df.to_dict(orient="records")
                
                return result
            elif OPENPYXL_AVAILABLE:
                # Use openpyxl to read Excel
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                
                # Get the sheet
                if sheet_name:
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook.active
                
                # Get headers if available
                headers = []
                if header:
                    for cell in sheet[1]:
                        headers.append(str(cell.value) if cell.value is not None else f"Column{cell.column}")
                else:
                    # Generate column names
                    for i in range(1, sheet.max_column + 1):
                        headers.append(f"Column{i}")
                
                # Read data
                result = []
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2 if header else 1, values_only=True)):
                    if row_idx < start_row - 1:
                        continue
                    
                    row_data = {}
                    for col_idx, value in enumerate(row):
                        if col_idx < len(headers):
                            row_data[headers[col_idx]] = value
                    
                    result.append(row_data)
                
                return result
            else:
                # Should not reach here due to check in can_read_excel()
                raise RuntimeError("Excel reading functionality is not available.")
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise
    
    def write_excel(self, file_path: str, data: List[Dict[str, Any]], 
                   sheet_name: str = "Sheet1", append: bool = False) -> None:
        """
        Write data to an Excel file
        
        Args:
            file_path: Path to Excel file
            data: List of row data dictionaries
            sheet_name: Sheet name to write
            append: Whether to append to existing file
        """
        if not self.can_write_excel():
            raise RuntimeError("Excel writing functionality is not available. Please install pandas and/or openpyxl.")
        
        try:
            if PANDAS_AVAILABLE:
                # Use pandas to write Excel
                df = pd.DataFrame(data)
                
                # Check if file exists and append mode is enabled
                if append and os.path.exists(file_path):
                    # Read existing file
                    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                        # Check if sheet already exists
                        reader = pd.ExcelFile(file_path)
                        if sheet_name in reader.sheet_names:
                            # Get existing sheet data
                            existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                            
                            # Concatenate with new data
                            result_df = pd.concat([existing_df, df], ignore_index=True)
                            
                            # Write back to file
                            result_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        else:
                            # Sheet doesn't exist, create new one
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    # Create new file or overwrite existing
                    df.to_excel(file_path, sheet_name=sheet_name, index=False)
            elif OPENPYXL_AVAILABLE:
                # Use openpyxl to write Excel
                
                # Get column headers
                headers = set()
                for row in data:
                    headers.update(row.keys())
                headers = list(headers)
                
                # Check if file exists and append mode is enabled
                if append and os.path.exists(file_path):
                    # Open existing workbook
                    workbook = openpyxl.load_workbook(file_path)
                    
                    # Check if sheet already exists
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        # Find last row
                        last_row = sheet.max_row
                        
                        # Append data
                        for row_idx, row_data in enumerate(data):
                            for col_idx, header in enumerate(headers):
                                sheet.cell(row=last_row + row_idx + 1, column=col_idx + 1, 
                                          value=row_data.get(header, None))
                    else:
                        # Create new sheet
                        sheet = workbook.create_sheet(sheet_name)
                        
                        # Write headers
                        for col_idx, header in enumerate(headers):
                            sheet.cell(row=1, column=col_idx + 1, value=header)
                        
                        # Write data
                        for row_idx, row_data in enumerate(data):
                            for col_idx, header in enumerate(headers):
                                sheet.cell(row=row_idx + 2, column=col_idx + 1, 
                                          value=row_data.get(header, None))
                    
                    # Save workbook
                    workbook.save(file_path)
                else:
                    # Create new workbook
                    workbook = openpyxl.Workbook()
                    
                    # Get active sheet and rename it
                    sheet = workbook.active
                    sheet.title = sheet_name
                    
                    # Write headers
                    for col_idx, header in enumerate(headers):
                        sheet.cell(row=1, column=col_idx + 1, value=header)
                    
                    # Write data
                    for row_idx, row_data in enumerate(data):
                        for col_idx, header in enumerate(headers):
                            sheet.cell(row=row_idx + 2, column=col_idx + 1, 
                                      value=row_data.get(header, None))
                    
                    # Save workbook
                    workbook.save(file_path)
            else:
                # Should not reach here due to check in can_write_excel()
                raise RuntimeError("Excel writing functionality is not available.")
            
            # Publish message
            self.message_bus.publish(MessageTypes.DATA_EXCEL_EXPORTED, {
                "file_path": file_path,
                "sheet_name": sheet_name,
                "row_count": len(data)
            })
            
            logger.info(f"Data exported to Excel file: {file_path} ({len(data)} rows)")
        except Exception as e:
            logger.error(f"Error writing Excel file: {e}")
            raise
    
    def map_columns(self, data: List[Dict[str, Any]], mappings: Dict[str, str]) -> List[Dict[str, Any]]:
        """
        Map columns in a dataset according to mappings
        
        Args:
            data: List of row data dictionaries
            mappings: Column mappings {source: target}
        
        Returns:
            list: Mapped data
        """
        result = []
        
        for row in data:
            mapped_row = {}
            
            for source, target in mappings.items():
                if source in row:
                    mapped_row[target] = row[source]
            
            result.append(mapped_row)
        
        return result
    
    def extract_column_values(self, data: List[Dict[str, Any]], column: str) -> List[Any]:
        """
        Extract values from a specific column
        
        Args:
            data: List of row data dictionaries
            column: Column name to extract
        
        Returns:
            list: Column values
        """
        result = []
        
        for row in data:
            if column in row:
                result.append(row[column])
        
        return result
    
    def filter_rows(self, data: List[Dict[str, Any]], filter_fn: callable) -> List[Dict[str, Any]]:
        """
        Filter rows using a filter function
        
        Args:
            data: List of row data dictionaries
            filter_fn: Filter function that takes a row and returns bool
        
        Returns:
            list: Filtered data
        """
        return [row for row in data if filter_fn(row)]